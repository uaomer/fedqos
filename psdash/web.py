# coding=utf-8
import csv
import ijson
import logging
import psutil
import socket
import uuid
import locale
import os
import openpyxl
import re
from psdash.node import LocalNode,RemoteNode
import sqlite3
import pygal
import networkx as nx
from networkx.drawing.nx_agraph import write_dot, graphviz_layout,to_agraph
import matplotlib.pyplot as plt  
import graphviz
import pydot 
from networkx import *
#import pygraphviz as pgv

import datetime
from datetime import datetime, timedelta
import arrow as arw 


from openpyxl import load_workbook   
from openpyxl.compat import range
from openpyxl.cell import cell


from flask import Flask, render_template, request, session, abort
from flask import json, jsonify, Response, Blueprint, current_app, g, flash, redirect, url_for
from flask_wtf import FlaskForm
from wtforms import Form,TextField,TextAreaField,validators,StringField,SubmitField, FileField, IntegerField, PasswordField, BooleanField
from werkzeug.local import LocalProxy
from werkzeug.utils import secure_filename 

from psdash.helpers import socket_families, socket_types
from fileinput import filename
from pyexcel.internal.sheets.row import Row
#from eventlet.support.dns.rdatatype import NULL
from pip._vendor.html5lib.html5parser import method_decorator_metaclass
from sqlalchemy.dialects.postgresql.base import CIDR
from flask.helpers import flash
#from keystoneclient.v3.contrib.trusts import Trust
from networkx.algorithms.traversal.breadth_first_search import bfs_edges
#from matplotlib.pyplot import arrow
from networkx.algorithms.shortest_paths.unweighted import predecessor
from networkx.algorithms.shortest_paths.generic import shortest_path_length
from wtforms.fields.simple import HiddenField
from wtforms.fields.core import SelectField, FloatField
from time import strftime
#from monasca_common.kafka_lib.consumer.base import FETCH_BUFFER_SIZE_BYTES


ALLOWED_EXTENSIONS = set(['txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'xlsx'])
ALLOWED_ANSWER_YES = set(['yes', 'Yes', 'YES', 'YEs','yES','yeS', 'Y', 'y', 'X', 'x'])
ALLOWED_ANSWER_NO = set(['no', 'No', 'NO', 'nO', 'N', 'n', 'X', 'x'])
ALLOWED_ANSWER_NA = set(['na', 'Na', 'NA', 'nA', 'N.A', 'N.A.', 'n.a.', 'N.a.','n.A.', 'not applicable','Not Applicable', 'not Applicable', 'Not applicable', 'X', 'x'])

logger = logging.getLogger('psdash.web')
webapp = Blueprint('psdash', __name__, static_folder='static')
conn = sqlite3.connect('db.sqlite3',detect_types=sqlite3.PARSE_DECLTYPES |sqlite3.PARSE_COLNAMES )

cur = conn.cursor()



def get_current_node():
    
    return current_app.psdash.get_node(g.node)

def get_current_service():
    return get_current_node().get_service()

current_node = LocalProxy(get_current_node)
current_service = LocalProxy(get_current_service)



def fromtimestamp(value, dateformat='%Y-%m-%d %H:%M:%S'):
    dt = datetime.fromtimestamp(int(value))
    return dt.strftime(dateformat)


@webapp.context_processor
def inject_nodes():        
    
  #  print "current node ",current_node
  #  print "get Node ", current_app.psdash.get_nodes()
    return {"current_node": current_node, "nodes": current_app.psdash.get_nodes()}


@webapp.context_processor
def inject_header_data():
    curtime = datetime.now()
    sysinfo = current_service.get_sysinfo()
    uptime = timedelta(seconds=sysinfo['uptime'])
    uptime = str(uptime).split('.')[0]
    return {
        'curtime':curtime,    
        'os': sysinfo['os'].decode('utf-8'),
        'hostname': sysinfo['hostname'].decode('utf-8'),
        'uptime': uptime
    }

@webapp.url_defaults
def add_node(endpoint, values):
    values.setdefault('node', g.node)

@webapp.before_request
def add_node():
    g.node = request.args.get('node', current_app.psdash.LOCAL_NODE)
    

@webapp.before_request
def check_access():
    if not current_node:
        return 'Unknown psdash node specified', 404

    allowed_remote_addrs = current_app.config.get('PSDASH_ALLOWED_REMOTE_ADDRESSES')
    if allowed_remote_addrs:
        if request.remote_addr not in allowed_remote_addrs:
            current_app.logger.info(
                'Returning 401 for client %s as address is not in allowed addresses.',
                request.remote_addr
            )
            current_app.logger.debug('Allowed addresses: %s', allowed_remote_addrs)
            return 'Access denied', 401

    username = current_app.config.get('PSDASH_AUTH_USERNAME')
    password = current_app.config.get('PSDASH_AUTH_PASSWORD')
    if username and password:
        auth = request.authorization
        if not auth or auth.username != username or auth.password != password:
            return Response(
                'Access deined',
                401,
                {'WWW-Authenticate': 'Basic realm="psDash login required"'}
            )

@webapp.before_request
def setup_client_id():
    if 'client_id' not in session:
        client_id = uuid.uuid4()
        current_app.logger.debug('Creating id for client: %s', client_id)
        session['client_id'] = client_id

@webapp.errorhandler(psutil.AccessDenied)
def access_denied(e):
    errmsg = 'Access denied to %s (pid %d).' % (e.name, e.pid)
    return render_template('error.html', error=errmsg), 401


@webapp.errorhandler(psutil.NoSuchProcess)
def access_denied(e):
    errmsg = 'No process with pid %d was found.' % e.pid
    return render_template('error.html', error=errmsg), 404

@webapp.route('/')
def index():
    if not session.get('logged_in'):
        return render_template('login.html')
    
    else:
        # process here for ziad
        
        sysinfo = current_service.get_sysinfo()
#         hostname = current_service.get_
        netifs = current_service.get_network_interfaces().values()
        netifs.sort(key=lambda x: x.get('bytes_sent'), reverse=True)
     #   print"Load average", sysinfo['load_avg']
        
#                         
        data = {
            'load_avg': sysinfo['load_avg'],
            'num_cpus': sysinfo['num_cpus'],
            'memory': current_service.get_memory(),
            'swap': current_service.get_swap_space(),
            'disks': current_service.get_disks(),
            'cpu': current_service.get_cpu(),
            'users': current_service.get_users(),
            'net_interfaces': netifs,
            'page': 'overview',
            'is_xhr': request.is_xhr 
           
        }
      
        return render_template('index.html', **data)


@webapp.route('/performance')
def performance():
    
    
    
    return render_template('performance.html', 
                           is_xhr= request.is_xhr)

@webapp.route('/perf')
def perf():
    
    sysinfo = current_service.get_sysinfo()
    netifs = current_service.get_network_interfaces().values()
    netifs.sort(key=lambda x: x.get('bytes_sent'), reverse=True)
    
    perf_data = current_service.get_myself()
    
   # print perf_data

    print perf_data 
    
#     for x in range(len(perf_data)-1): 
#         print perf_data[x]
#     
# ur.execute('select count(id) from caiqanswer where cloud_id=:1 and choice_id=:2 and cgroup_id=:3', (cid,'1',cgroup_id))
#     cur.execute("select id, cname from cprofile where cendpoint=:1",[(perf_data[0][1])])
#     whois = cur.fetchone()
#     print whois[0]    
#     
    cur.execute("select id from performance where timemilli=:1 and cname=:2 ",( perf_data[0][0], perf_data[0][1]))
       
    fetch_data = cur.fetchall()
     
    if not fetch_data:
        for each_perf in perf_data:  
            print each_perf
            cur.execute("insert into performance (timemilli,cname,projectid, taskid, intime,outtime,ifcsize,objsize, message) values (?,?,?,?,?,?,?,?,?)", (each_perf[0],each_perf[1], each_perf[2],each_perf[3],each_perf[4],each_perf[5],each_perf[6], each_perf[7],each_perf[8]  ))
            conn.commit()
            
    cur.execute("select performance.cname, SUM(performance.outtime - performance.intime), SUM(performance.objsize) from performance where performance.projectid=:1 and message LIKE 'worker%'", ([perf_data[0][2]])) 
    fetch_data = cur.fetchone()
    print fetch_data
    avg_perf = round(float(fetch_data[2])/fetch_data[1],5) 
    print "average performance",avg_perf, "objects per seconds"
    
    cur.execute("update cprofile set pvalue=:1 where cprofile.cendpoint=:2", ( avg_perf, fetch_data[0]))
    conn.commit()

    data = {
        'load_avg': sysinfo['load_avg'],
        'num_cpus': sysinfo['num_cpus'],
        'memory': current_service.get_memory(),
        'swap': current_service.get_swap_space(),
        'disks': current_service.get_disks(),
        'cpu': current_service.get_cpu(),
        'users': current_service.get_users(),
        'net_interfaces': netifs,
        'page': 'overview',
        'is_xhr': request.is_xhr
        }
  
    return render_template('perf.html',performance=perf_data, **data)

@webapp.route('/processes', defaults={'sort': 'cpu_percent', 'order': 'desc', 'filter': 'user'})
@webapp.route('/processes/<string:sort>')
@webapp.route('/processes/<string:sort>/<string:order>')
@webapp.route('/processes/<string:sort>/<string:order>/<string:filter>')
def processes(sort='pid', order='asc', filter='user'):
    procs = current_service.get_process_list()
    num_procs = len(procs)

    user_procs = [p for p in procs if p['user'] != 'root']
    num_user_procs = len(user_procs)
    if filter == 'user':
        procs = user_procs

    procs.sort(
        key=lambda x: x.get(sort),
        reverse=True if order != 'asc' else False
    )

    
    return render_template(
        'processes.html',
        processes=procs,
        sort=sort,
        order=order,
        filter=filter,
        num_procs=num_procs,
        num_user_procs=num_user_procs,
        page='processes',
        is_xhr=request.is_xhr
    )


@webapp.route('/process/<int:pid>', defaults={'section': 'overview'})
@webapp.route('/process/<int:pid>/<string:section>')
def process(pid, section):
    valid_sections = [
        'overview',
        'threads',
        'files',
        'connections',
        'memory',
        'environment',
        'children',
        'limits'
    ]

    if section not in valid_sections:
        errmsg = 'Invalid subsection when trying to view process %d' % pid
        return render_template('error.html', error=errmsg), 404

    context = {
        'process': current_service.get_process(pid),
        'section': section,
        'page': 'processes',
        'is_xhr': request.is_xhr
    }

    if section == 'environment':
        penviron = current_service.get_process_environment(pid)

        whitelist = current_app.config.get('PSDASH_ENVIRON_WHITELIST')
        if whitelist:
            penviron = dict((k, v if k in whitelist else '*hidden by whitelist*') 
                             for k, v in penviron.iteritems())

        context['process_environ'] = penviron
    elif section == 'threads':
        context['threads'] = current_service.get_process_threads(pid)
    elif section == 'files':
        context['files'] = current_service.get_process_open_files(pid)
    elif section == 'connections':
        context['connections'] = current_service.get_process_connections(pid)
    elif section == 'memory':
        context['memory_maps'] = current_service.get_process_memory_maps(pid)
    elif section == 'children':
        context['children'] = current_service.get_process_children(pid)
    elif section == 'limits':
        context['limits'] = current_service.get_process_limits(pid)

    return render_template(
        'process/%s.html' % section,
        **context
    )


@webapp.route('/network')
def view_networks():
    netifs = current_service.get_network_interfaces().values()
    netifs.sort(key=lambda x: x.get('bytes_sent'), reverse=True)

    # {'key', 'default_value'}
    # An empty string means that no filtering will take place on that key
    form_keys = {
        'pid': '', 
        'family': socket_families[socket.AF_INET],
        'type': socket_types[socket.SOCK_STREAM],
        'state': 'LISTEN'
    }

    form_values = dict((k, request.args.get(k, default_val)) for k, default_val in form_keys.iteritems())

    for k in ('local_addr', 'remote_addr'):
        val = request.args.get(k, '')
        if ':' in val:
            host, port = val.rsplit(':', 1)
            form_values[k + '_host'] = host
            form_values[k + '_port'] = int(port)
        elif val:
            form_values[k + '_host'] = val

    conns = current_service.get_connections(form_values)
    conns.sort(key=lambda x: x['state'])

    states = [
        'ESTABLISHED', 'SYN_SENT', 'SYN_RECV',
        'FIN_WAIT1', 'FIN_WAIT2', 'TIME_WAIT',
        'CLOSE', 'CLOSE_WAIT', 'LAST_ACK',
        'LISTEN', 'CLOSING', 'NONE'
    ]

    return render_template(
        'network.html',
        page='network',
        network_interfaces=netifs,
        connections=conns,
        socket_families=socket_families,
        socket_types=socket_types,
        states=states,
        is_xhr=request.is_xhr,
        num_conns=len(conns),
        **form_values
    )

@webapp.route('/disks')
def view_disks():
    disks = current_service.get_disks(all_partitions=True)
    io_counters = current_service.get_disks_counters().items()
    io_counters.sort(key=lambda x: x[1]['read_count'], reverse=True)
    return render_template(
        'disks.html',
        page='disks',
        disks=disks,
        io_counters=io_counters,
        is_xhr=request.is_xhr
    )

@webapp.route('/logs')
def view_logs():
    available_logs = current_service.get_logs()
    available_logs.sort(cmp=lambda x1, x2: locale.strcoll(x1['path'], x2['path']))

    return render_template(
        'logs.html',
        page='logs',
        logs=available_logs,
        is_xhr=request.is_xhr
    )


@webapp.route('/log')
def view_log():
    filename = request.args['filename']
    seek_tail = request.args.get('seek_tail', '1') != '0'
    session_key = session.get('client_id')

    try:
        content = current_service.read_log(filename, session_key=session_key, seek_tail=seek_tail)
    except KeyError:
        error_msg = 'File not found. Only files passed through args are allowed.'
        if request.is_xhr:
            return error_msg
        return render_template('error.html', error=error_msg), 404

    if request.is_xhr:
        return content

    return render_template('log.html', content=content, filename=filename)


@webapp.route('/log/search')
def search_log():
    filename = request.args['filename']
    query_text = request.args['text']
    session_key = session.get('client_id')

    try:
        data = current_service.search_log(filename, query_text, session_key=session_key)
        return jsonify(data)
    except KeyError:
        return 'Could not find log file with given filename', 404


@webapp.route('/register')
def register_node():
    name = request.args['name']
    port = request.args['port']
    host = request.remote_addr

    current_app.psdash.register_node(name, host, port)
    return jsonify({'status': 'OK'})


# handling forms the flasky way 

class SignupForm(Form):
       
 
    username = TextField('Username', validators=[validators.Length(min=4, max=20)])
    password = PasswordField('New Password', validators=[
        validators.Required(),
        validators.EqualTo('confirm', message='Passwords must match')
    ])
    confirm = PasswordField('Repeat Password')
    accept_tos = BooleanField('I accept the Terms of Service and Privacy Notice (updated Aug 22, 2017)', validators=[validators.Required()])

class RegisterForm(Form):
    
    login_id = HiddenField('Login ID:', validators=[validators.required()])
    username = TextField('User Name:', validators=[validators.required()])
    cname = TextField('Cloud Name:', validators=[validators.required()])
    cmeta = TextField('Metatext:', validators=[validators.required()])    
    cendpoint = TextField('Endpoint URL:', validators=[validators.required()])
    cinfo = TextField('Description:', validators=[validators.required()])
        
class AddResource(Form):
     
    transaction_id = TextField('Transaction ID')
    hcloud_id = TextField('HCloud ID')
    hcloud_name = TextField('Home Cloud')
    cloud_id = HiddenField('Cloud ID',[validators.Required()])
    cname = TextField('Cloud Name:',[validators.Required()])
    
    rtype = SelectField(u'Resource Type', choices=[('1','IaaS')])
    #resource = SelectField(u'Resource', choices=[('1','Compute'),('2','Storage'),('3','Network')])
    resource = SelectField(u'Resource', coerce=int)
    #rthreshold = FloatField('Trust threshold', validators = [validators.Required() ])
    rvalue = FloatField('Resource Value', validators = [validators.Required() ])
    uom =  TextField('Unit', [validators.Required()])
    rqty = IntegerField('Quantity', [validators.Required()] )
    lstart =  TextField('Time to start', [validators.Required()])
    expiresat =  TextField('Expires at :', [validators.Required()])
    accept_tos = BooleanField('I accept the Terms of Service and Privacy Notice (updated Sept 22, 2017)', [validators.Required()])
    
    def set_choices(self):
        cur.execute("select id, rname from resources")
        resource_choices= cur.fetchall()
        #printresource_choices
        self.resource.choices = resource_choices 

    
@webapp.route("/signup", methods=['GET', 'POST'])
def signup():
    form = SignupForm(request.form)
    #printform.errors
    
    if request.method == 'POST':
        
        username = slugify(request.form['username'],1) # convert to lower here -- 
        password = request.form['password']
        confirm = request.form['confirm']
        
        if form.validate():
            #printusername
            cur.execute('select * from login where username=?', [(username)])
            rows = cur.fetchall()
            if rows: 
                flash("That username is already taken, please choose another")
                return render_template('signup.html', form=form, is_xhr=request.is_xhr)
            else:
                cur.execute('INSERT Into login(username,password,access_level) values (?,?,?)', (username, password, 4))
                conn.commit()
                cur.execute('select * from login where username = :1 ', [(username)])
                whois = cur.fetchone()
                #print"This is user id ", whois[0]

                if whois:
                    #printwhois[0]
                    #printwhois[1]
                    session['login_id'] = whois[0]
                    session['username'] = whois[1]
                    session['access_level'] = whois[2]    
                
#                return redirect('/upload') # this will be register a cloud. 
                    #return render_template('cregister.html')
                return redirect("/cregister")
        else:
            flash('All the form fields are required. ')
 
    return render_template('signup.html', form=form,is_xhr=request.is_xhr )

@webapp.route("/cregister", methods=['GET', 'POST'])
def register():
    
    #print"gotcha0"
    form = RegisterForm(request.form) # create the registration form 
    ##printform.errors
    #print"gotcha1"
    
    if request.method == 'POST': # receive the value if request is posted 
        
        login_id = request.form['login_id']
        username = request.form['username']
        cname= request.form['cname']
        cmeta = request.form['cmeta']
        cendpoint = request.form['cendpoint']
        cinfo = request.form['cinfo']
        #print"gotcha2"
        if form.validate():
            #print"gotcha3"
            #printlogin_id 
            cur.execute('select id,cname from cprofile where login_id=?', [(login_id)])
            row = cur.fetchone() # if there is any other cloud with this login iD 
            if row:
                #print"inside cur.rowcount" 
                flash("A cloud provider is already registered with this Login ID.")
                #print"A cloud provider is already registered with this Login ID."
                #printrow[0]
                cur.execute("select * from caiqanswer where cloud_id=:1 ", [(row[0])])
                rows = cur.fetchall()
                if len(rows) < 1: 

                    flash("Complete the signup process by uploading your caiq")
                    session['cloud_id'] = row[0]
                    session['cloud_name'] = row[1]
                    
                    return render_template('upload.html')
                      
                else: # register another cloud 
                    #print"Register another cloud with another name "
                    return render_template('signup.html', form=form, is_xhr=request.is_xhr)
                    
            else:
                cur.execute("INSERT INTO cprofile(cname,cmeta,cendpoint,cinfo, login_id) VALUES (?,?,?,?,?)", (cname,cmeta,cendpoint,cinfo, login_id)  ) 
                conn.commit()
                flash('New cloud added to database successfully.') #Display a message to end user at front end.
                #print'New cloud added to database successfully.' #Display a message to end user at front end.
                cur.execute("select id, cname from cprofile where cname=?" , [(cname)]  )
                whois = cur.fetchone()

                if whois:                   
                    session['cloud_id'] = whois[0]
                    session['cloud_name'] = whois[1]
                    
                    return redirect('/upload') # this will be register a cloud. 
        else:
            flash('All the form fields are required. ')
 
    return render_template('cregister.html', form=form,is_xhr=request.is_xhr )


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
           
@webapp.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # check if the post request has the file part
        
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        
        ufile = request.files['file']
        # if user does not select file, browser also
        # submit a empty part without filename
        if ufile.filename == '':
            flash('No selected file')
            #print('No selected file')
            return redirect(request.url)
        if not allowed_file(ufile.filename):
            #print('File type not permitted')
            return redirect(request.url)
        
        if ufile and allowed_file(ufile.filename):
            filename = secure_filename(ufile.filename)            
            svar_cname = session.get('cloud_name', None)
            svar_cid = session.get('cloud_id', None)
            ufile.save(os.path.join('/tmp/',svar_cname))
            
            wb = load_workbook(ufile)
            sheets = wb.get_sheet_names()
            #printsheets
            
            for sheet in sheets:
                ws = wb[sheet] 
                #printws 
                
                maxm_row = ws.max_row
                #printmaxm_row
                maxm_col = ws.max_column
                #printmaxm_col
               
                for x in range (2,maxm_row+1):
                    #for y in range (1,maxm_col+1):
                    cgroup_id=ws.cell(row=x,column=1).value  
                    cont_id=ws.cell(row=x,column=2).value
                    ayes = ws.cell(row=x,column=3).value
                    ano = ws.cell(row=x,column=4).value
                    ana = ws.cell(row=x,column=5).value
                    
                    if ayes in ALLOWED_ANSWER_YES:
                            ans = '1'
                    elif ano in ALLOWED_ANSWER_NO:
                        ans = '2'      
                    elif ana in ALLOWED_ANSWER_NA: 
                        ans = '3'
                    else: 
                        ans = '4'
                    
                    cur.execute("INSERT INTO caiqanswer (control_id,choice_id, cloud_id,cgroup_id) VALUES (?,?,?,?)", (cont_id,ans,svar_cid,cgroup_id)  )
                    conn.commit()
                #print'success!!!!'
            flash('File upload success!!')
            
            cur.execute('select short_code from caiqcgroup ')
            group_dict = cur.fetchall()
            c_trust_eval=[]
            for group in group_dict: 
                #printgroup[0]
                c_trust_eval.append (cgroup_score(svar_cid, group[0])) # Evaluating trust based on caiq for each control group 
            
            caiq_trust_score(svar_cid)
            
            #print"This is where to look Login ID=", session['login_id']
            cur.execute('update login set access_level=2 where login.id=?', [(session['login_id'])])
            conn.commit()
                
            return render_template("login.html")
    return render_template('upload.html')

def check_login(): 
    if not session.get('logged_in')==True:
        #print"i am alive"
        return redirect('/login')
        #print"i am alive2"
    else:
        #print"i am alive3"
        
        return redirect( request.referrer)
@webapp.route("/logout")
def logout():
    session['logged_in'] = False
    #session.clear()
    
    return redirect('/login')
        
@webapp.route("/login", methods=['GET','POST'])
def web_login():
    
    if request.method == 'POST': 
        username = request.form['username']
        password = request.form['password']
    
        cur.execute('select * from login where username = lower(:1) and password=:2', (username, password))
        whois = cur.fetchone()
        #printwhois
        if whois: 
            session['logged_in'] = True
            session['login_id'] = whois[0]
            session['username'] = whois[1]
            
            session['access_level'] = whois[3]
            if whois[3] == 4: 
                
                return redirect("/cregister")
            else:
                reg_clouds = []
                cur.execute('select id,cname from cprofile where login_id =?', [(session['login_id'])])
                cloud_ids =  cur.fetchall()
                for clouds in cloud_ids: 
                    reg_clouds.append(clouds)
                session['reg_clouds'] = reg_clouds
                
                return redirect("/profiles")
        else:
            flash('wrong password!')
            return render_template('login.html') 
    return render_template('login.html')



def slugify(text, lower=1):
    if lower == 1:
        text = text.strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text



# @webapp.route('/profiles', defaults={'sort': 'avg_e', 'order': 'desc'})
# @webapp.route('/profiles/<string:sort>')
# @webapp.route('/profiles/<string:sort>/<string:order>')
# def profiles(sort='id', order='asc'):
#     
#     #check_login()
#     query1= """select id,cname,cendpoint,
#             strftime('%s','now','localtime') - strftime('%s',lastseen) AS 'timesince',
#             strftime('%Y-%m-%d', lastseen),
#             avg_e, 
#             strftime('%H:%M:%S', lastseen),
#             avg_t,avg_c,avg_e,avg_f,avg_b,avg_d,avg_u,avg_w, pvalue  
#             from cprofile  order by """ 
#     query2 =  sort + ' ' + order
#     ##printquery1+query2
#     cur.execute (query1+query2)
#     all_profiles = cur.fetchall()
#     
#     
#     
#     if (session['reg_clouds']):
#         cid = (session['reg_clouds'][0][0])
#         importance = profile_detail(cid)[22]
#     else: 
#         importance = 1 
#   #  print importance 
#          
#     return render_template('profiles.html', importance=importance, profiles=all_profiles, sort=sort, order=order, page='profiles', is_xhr=request.is_xhr)

@webapp.route('/profiles', defaults={'sort': 'avg_e', 'order': 'desc'})
@webapp.route('/profiles/<string:sort>')
@webapp.route('/profiles/<string:sort>/<string:order>')
@webapp.route('/profiles/<string:sort>/<string:order>/<int:importance>')
def profiles(sort='id', order='asc', importance=1):
    
    #check_login()
    query1= """select id,cname,cendpoint,
            (select discipline.dname from discipline where discipline.id=cprofile.cmeta),
            strftime('%s','now','localtime') - strftime('%s',lastseen) AS 'timesince',
            strftime('%Y-%m-%d', lastseen), 
            strftime('%H:%M:%S', lastseen), 
            avg_t,avg_c,avg_e,avg_f,avg_b,avg_d,avg_u,avg_w, pvalue  
            from cprofile  order by """ 
    query2 =  sort + ' ' + order
    query3 = 'group by cinfo' 
    
    ##printquery1+query2
    cur.execute (query1+query2)
    all_profiles = cur.fetchall()
    
    print "I got importance from URL", importance
    
#     if (session['reg_clouds']):
#         cid = (session['reg_clouds'][0][0])
#         importance = profile_detail(cid)[22]
#     else: 
#         importance = 1 
#   #  print importance 
        
    return render_template('profiles.html', importance=importance, profiles=all_profiles, sort=sort, order=order, page='profiles', is_xhr=request.is_xhr)


@webapp.route('/profile/<int:cid>', defaults={'section': 'overview'})
@webapp.route('/profile/<string:cid>', defaults={'section': 'overview'})
@webapp.route('/profile/<int:cid>/<string:section>')
def profile(cid, section):
    ##print"This is profile function", cid
    valid_sections = [
        'overview',
        'assessment',
        'resources',
        'trust',
        'graph',
        ]

    if section not in valid_sections:
        errmsg = 'Invalid subsection when trying to view detail' 
        return render_template('error.html', error=errmsg), 404
    
    
    # trust evaluation starts here - Note: this must move to file upload section
    c_trust_eval= []
    select_profile = profile_detail(cid) # profile details 
    #print"This is cloud ", select_profile
    
    caiq_assessment_detail = caiq_detail(cid) # caiq details 
    #print"This is CAIQ assessment", caiq_assessment_detail

    #cur.execute ('select ctrustinfo.*,caiqcgroup.* from ctrustinfo INNER Join caiqcgroup on ctrustinfo.cgroup_id=caiqcgroup.id and ctrustinfo.cloud_id=?', [(cid)])
    cur.execute ('''select caiqcgroup.group_name,caiqcgroup.nofquestion,ctrustinfo.count_na, 
                        (caiqcgroup.nofquestion-ctrustinfo.count_na) AS "Total Applicable", ctrustinfo.count_yes,
                        ctrustinfo.count_no, ctrustinfo.count_un,ctrustinfo.caiq_t,ctrustinfo.caiq_c,
                        ctrustinfo.caiq_f,ctrustinfo.caiq_e, 
                        ctrustinfo.caiq_b,ctrustinfo.caiq_d,ctrustinfo.caiq_u,ctrustinfo.caiq_w  
                        from ctrustinfo 
                        INNER Join caiqcgroup on ctrustinfo.cgroup_id=caiqcgroup.id and 
                        ctrustinfo.cloud_id=?''', [(cid)])
    ctrust= cur.fetchall()
    
    cur.execute('''select all_na, (295-all_na) AS "Total Applicable", 
                        all_yes,all_no,all_un,avg_t,avg_c,avg_f,avg_e,avg_b, avg_d,avg_u,avg_w  
                        from cprofile where id=?'''  ,[(cid)])
    
    misc_data = cur.fetchone()
    #print"MISC DATA lOOK",misc_data
    
#     view_trust_detail = cur.fetchall()
#     #print"This is trust detail", view_trust_detail
#     misc_data = 1,2,3,4,5,6,7,8,9,10 # just a jugad
#     
    graph_data = draw_graph(cid)
    
    context = {
        'profile': select_profile,
        'caiq': caiq_assessment_detail,
        'ctrust':ctrust,
        'misc_data': misc_data,
        'graph_data': graph_data,
        'section': section,
        'page': 'profiles',
        'is_xhr': request.is_xhr
    }
    
    if section == 'assessment':

        context['assessment'] = 'This is profile assessment'
       
        
    elif section == 'resources':
        context['resources'] = 'This is resources panel'
       
    elif section == 'trust':

        context['trust'] = 'This is trust management'
       # #print"This is full context= ", context

    elif section == 'graph':
    
        context['graph'] = 'This is graphical representation of trust'
      
 
    return render_template(
        'profile/%s.html' % section,
        **context
    )

    
def profile_detail(cid):
    
    cur.execute ("select * from cprofile where id=?", [cid])
    
    #cur.execute("select cprofile.*, TempTable.* from TempTable INNER Join cprofile on TempTable.cloud_id = cprofile.id and TempTable.cloud_id=?",[cid])
    whois = cur.fetchone()  
    
    if not whois:
        errmsg = 'Invalid profile when trying to view detail' 
        return render_template('error.html', error=errmsg), 404
    return whois

def caiq_detail(cid):
    
    cur.execute('select caiqanswer.cloud_id, caiqanswer.control_id, caiqquestion.q_text, caiqanswer.choice_id, caiqchoice.choice_text  from ((caiqanswer LEFT Join caiqquestion on caiqanswer.control_id=caiqquestion.id) inner join caiqchoice on caiqanswer.choice_id=caiqchoice.id) where cloud_id=?', [cid]) 
      
    caiq = cur.fetchall()
    
    if not caiq:
        #print'Nothing found'
        errmsg = 'Invalid profile when trying to view detail' 
        return render_template('error.html', error=errmsg), 404
    
    return caiq

def cgroup_score(cid,cgroup):
    
    #print"this is caiq_funtion. I got CID=", cid, 'and cgroup=',cgroup
    cur.execute('select id,group_name,nofquestion from caiqcgroup where short_code like :1', ([cgroup]))
    result = cur.fetchone()
    cgroup_id = result[0]
    cgroup_name = result[1]
    count_total = result[2]
    #print"This is group:", cgroup_name, "with ID: ", cgroup_id
    #print"Total Question=", count_total
    
    cur.execute('select count(id) from caiqanswer where cloud_id=:1 and choice_id=:2 and cgroup_id=:3', (cid,'1',cgroup_id))
    result = cur.fetchone()
    count_yes = result[0]
    #print'Yes=     ', count_yes
    
    cur.execute('select count(id) from caiqanswer where cloud_id=:1 and choice_id=:2 and cgroup_id=:3', (cid,'2',cgroup_id)) 
    result = cur.fetchone()
    count_no = result[0]
    #print'No=      ', count_no
    
    cur.execute('select count(id) from caiqanswer where cloud_id=:1 and choice_id=:2 and cgroup_id=:3', (cid,'3',cgroup_id)) 
    result = cur.fetchone()
    count_na = result[0]
    #print'NA=      ', count_na
    
    cur.execute('select count(id) from caiqanswer where cloud_id=:1 and choice_id=:2 and cgroup_id=:3', (cid,'4',cgroup_id)) 
    result = cur.fetchone()
    count_un = result[0]
    #print'Unknown= ', count_un
    
    #print'----------------------------'
    
    total_applicable = (count_no + count_un + count_yes)
    #print"total applicable", total_applicable
    
#     sum_y_no = float(count_yes[0]+ count_no[0])
#     #printsum_y_no
    
    if  (count_yes + count_no) == 0: 
        caiq_t=0
    else: 
        caiq_t = round(count_yes / (float(count_yes + count_no)),5)
        #print"CAIQ T= ", caiq_t
    
#     temp1= ta_assessment * (yes_assessment + no_assessment)    
    
    if total_applicable==0:
        caiq_c = 1
        #print"Total applicable was Zero so CAIQ_C= ",caiq_c 
    else:
        temp1 = total_applicable * (count_yes+count_no)
        #print"temp1 =", temp1
        
    #     temp2 = 2*(ta_assessment - (yes_assessment+no_assessment))
        temp2 = 2*(total_applicable - (count_yes+count_no))
        #print"temp2=", temp2
    #     trust_c= temp1 / (temp2+temp1)
        caiq_c = round(temp1 / float(temp2+temp1),5)
        #print"-------------------------------"
        #print"CAIQ_C= ", caiq_c

#     trust_f= 0.99
    caiq_f = 0.99
    
#     trust_e = trust_t * trust_c + (1-trust_c)*trust_f
    caiq_e = caiq_t * caiq_c + (1-caiq_c)*caiq_f
    #print"CAIQ E Score" , caiq_e
    
    #caiq_t = round(count_yes / (float(count_yes + count_no)),4)
    
    # Direct calculation based on Josang'theory. Differs from Ries paper and creates problems due to 
    # addition of N in Ries' theory. 
      
#     caiq_b = round(count_yes / float(count_yes + count_no + 2),4)
#     #printcaiq_b
#     caiq_d = round(count_no / float(count_yes + count_no + 2),4)
#     #printcaiq_d
#     caiq_u = round(2 / float(count_yes + count_no + 2),4)
#     #printcaiq_u

    caiq_b = caiq_t * caiq_c # conversion 
    caiq_b = round(caiq_b,5)
    
    caiq_d = (1-caiq_t) * caiq_c
    caiq_d = round(caiq_d,5)
    
    caiq_u = 1-caiq_c 
    caiq_u = round(caiq_u,5)
    
    caiq_a = caiq_f
    caiq_a = round(caiq_a,5)
    
    caiq_w = caiq_b+caiq_u*caiq_a
    caiq_w = round(caiq_w,5)
    
    #printcaiq_a,caiq_b,caiq_c, caiq_d, caiq_e, caiq_f, caiq_t, caiq_u, caiq_w

            
    cur.execute("""INSERT INTO ctrustinfo(cgroup_id,cloud_id,caiq_t,caiq_c,caiq_e,caiq_f,caiq_b,caiq_d,caiq_u,
        count_yes,count_no,count_na,count_un,caiq_w) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (cgroup_id,cid,caiq_t,caiq_c,caiq_e,caiq_f,caiq_b,caiq_d, caiq_u,count_yes,count_no,count_na,
         count_un,caiq_w)) 
    
    conn.commit()
         
    ##printtotal_applicable   
    return {'count_yes':count_yes,'count_no':count_no,'count_na':count_na,
            'count_un':count_un,'total_applicable': total_applicable, 
            'caiq_t': round(caiq_t,5), 'caiq_c':round(caiq_c,5),
            'caiq_e':round(caiq_e,5),'caiq_f':round(caiq_f,5),
            'cgroup':cgroup_name,'count_total':count_total
            } 
    
@webapp.route('/edit', methods=['GET', 'POST'] )
#@webapp.route('/edit/<string:cid>' )
def edit_profile():
    
    if request.method == 'POST':
        
        if request.form['faction'] == 'new': 
            return redirect('/signup')
        
        if request.form['faction'] == 'composite':
            if 'cart' not in session:
                flash ("Nothing to do")
                return redirect("/profiles")
            else: 
                return redirect("/composite")
        
        if request.form['faction'] == 'Refresh Cart':
            if 'cart' not in session:
                flash ("Nothing to do")
                #return redirect("/profiles")
                return redirect(request.referrer)

            else: 
                cloud_list= []
                for cid in session.pop('cart', []):
                    cloud_list.append(cid)
                flash ("Cart is refreshed")
                #return redirect("/profiles")
                return redirect(request.referrer)

            
        if request.form['faction'] == 'compare':  
            if 'cart' not in session:
                flash ("Nothing to compare")
                #return redirect("/profiles")
                
                return redirect(request.referrer)
            else: 
                return redirect("/compare")
            
#         if request.form['imp'] and request.form['faction'] == 'importance':
#             imp = request.form['imp']
#             faction = request.form['faction']
#             cid = (session['reg_clouds'][0][0]) 
#             
#             cur.execute('update cprofile set tthreshold=:1 where id=:2' , (imp, cid))
#             conn.commit()
#             
#             return redirect("/profiles")
        
        if request.form['cid'] and request.form['faction'] == 'delete':
            cid = request.form['cid']
            faction = request.form['faction']
            
            #print'For delete=', faction, cid
                     
            cur.execute( 'delete from cprofile where id IN (:1)',([cid]))
            conn.commit()
            #msg='Profile deleted successfully'
            flash('Profile deleted successfully')
            print("affected rows = {}".format(cur.rowcount))
                 
    #return redirect('/profiles')
    return redirect(request.referrer)


  
@webapp.route('/compare', defaults={'section': 'overview'})
@webapp.route('/compare/<string:section>')
def compare_trust(section):
    ##print"This is profile function", cid
#     test_run = current_service.run_script()
#     print "Running command ", test_run
#     
    
    valid_sections = [
        'overview',
        'objective',
        'subjective',
        'graph',
        ]

    if section not in valid_sections:
        errmsg = 'Invalid subsection when trying to view detail' 
        return render_template('error.html', error=errmsg), 404
    
    #cloud_list=[]
    
    #print"session cart ", session['cart']
     
#     for cid in session.pop('cart', []):
#         cloud_list.append(cid)
    cloud_list = (session['cart'])
    #printcloud_list
    #printlen(cloud_list)
    if len(cloud_list)==1: 
        flash("Can't compare single item") 
        return redirect("/profiles") 
    
    fetch_data = caiq_compare(cloud_list)
    result_compare = fetch_data.get('result_compare')
    cloud_detail = fetch_data.get('cloud_detail')
    parsed_result = fetch_data.get('parsed_result')
    graph_data = draw_graph1(cloud_list)
    
    context = {
        'parsed_result': parsed_result,
        'cloud_detail': cloud_detail,
        'cloud_list':cloud_list,
        'result_compare':result_compare,
        'graph_data': graph_data,
        'section': section,
        'page': 'compare',
        'is_xhr': request.is_xhr
    }
    
    if section == 'objective':
        context['objective'] = 'This is objective trust comparison' 
    elif section == 'subjective':
        context['subjective'] = 'This is subjective trust comparison'
    elif section == 'graph':
        context['graph'] = 'This is graphical representation of trust'
 
    return render_template(
        'compare/%s.html' % section,
        **context
    )

@webapp.route('/add_to_compare/<string:cid>' )
def add_to_compare(cid):
    if 'cart' in session:
        if len(session['cart']) == 4:
            flash('Cart Full..Cant add more')
            exit
        else:
            session['cart'].append(cid)
            flash(len(session['cart']))
    else:
        session['cart'] = [cid]
        flash(len(session['cart']))
    return redirect( request.referrer)    

@webapp.route('/importance/<string:imp>' )
def set_importance(imp):
    
   # print "This is the new importtance ", imp
    return redirect( request.referrer)    


def draw_graph(cid): # dont delete now..delete when  it will definitly merge into the new function 
    
    #Plotting all values for trust as graph    
    #print"This is graph func ", cid
    graph = pygal.Line(range=(0,1.2 )) 
    graph.title = ' Trust Evaluation for each control '
    x_label = []
    
    cur.execute('select short_code from caiqcgroup ORDER BY id')
    groups = cur.fetchall()
    for group in groups: # graph labels 
        x_label.append(group[0])    
    graph.x_labels = x_label

    #select ctrustinfo.caiq_e,ctrustinfo.cgroup_id,caiqcgroup.group_name,caiqcgroup.short_code from ctrustinfo  inner join caiqcgroup on ctrustinfo.cgroup_id=caiqcgroup.id and ctrustinfo.cloud_id=51  
    mark_list =[]
    cur.execute('select ctrustinfo.caiq_e,ctrustinfo.cgroup_id,caiqcgroup.group_name,caiqcgroup.short_code from ctrustinfo inner join caiqcgroup on ctrustinfo.cgroup_id=caiqcgroup.id and ctrustinfo.cloud_id=? ORDER BY cgroup_id', [(cid)])
    c_trust_eval = cur.fetchall()
    for x in c_trust_eval:
        mark_list.append(x[0])
    avg_caiq_e = sum(mark_list)/16
    #printmark_list, avg_caiq_e
    
    graph.add('CAIQ E-Score',mark_list)
    
    list_avg_e = [] # plotting average score as blue line against each dot .--.--.--.--.--. 
    for x in range(len(mark_list)):         
        list_avg_e.append(avg_caiq_e)          
      
    graph.add("CAIQ Avg Escore", list_avg_e, show_dots=False)
    graph_data = graph.render_data_uri()    # drawing the graph
    
    #print"graph was a success"
# 
    return graph_data


def draw_graph1(cloud_list):
        
    #print"This is the new graph func ", cloud_list
    
    graph = pygal.Line(range=(0,1.2 )) 
    graph.title = 'Comparison of Trust Evaluation for all selected clouds '
    x_label = []
     
    cur.execute('select short_code from caiqcgroup ORDER BY id')
    groups = cur.fetchall()
    for group in groups: # graph labels 
        x_label.append(group[0])    
    graph.x_labels = x_label
 
    
    for cid in cloud_list:
        cur.execute('select cname from cprofile where id=:1',[(cid)])
        cloud_name = cur.fetchone()
        
         
        full_query= '''select ctrustinfo.caiq_e,ctrustinfo.cgroup_id,
                    caiqcgroup.group_name,caiqcgroup.short_code 
                    from ctrustinfo inner join caiqcgroup on ctrustinfo.cgroup_id=caiqcgroup.id 
                    and ctrustinfo.cloud_id= (%s) ORDER BY cgroup_id''' % cid 
 
        mark_list =[]
        cur.execute(full_query)
        c_trust_eval = cur.fetchall()
        
        for x in c_trust_eval:
            mark_list.append(x[0])
        graph.add('(%s)' % cloud_name  , mark_list)

    graph_data = graph.render_data_uri()    # drawing the graph
    
    return graph_data


def caiq_trust_score(cid): # its one time function for every profile. never call twice 
    
    cur.execute('select count(id) from caiqanswer where caiqanswer.cloud_id=:1 and caiqanswer.choice_id=1', [(cid)])
    sum_all_yes = cur.fetchone()
     
    cur.execute('select count(id) from caiqanswer where caiqanswer.cloud_id=:1 and caiqanswer.choice_id=2', [(cid)])
    sum_all_no = cur.fetchone()
    
    cur.execute('select count(id) from caiqanswer where caiqanswer.cloud_id=:1 and caiqanswer.choice_id=3', [(cid)])
    sum_all_na = cur.fetchone()
    
    cur.execute('select count(id) from caiqanswer where caiqanswer.cloud_id=:1 and caiqanswer.choice_id=4', [(cid)])
    sum_all_un = cur.fetchone()
    
    
    avg_caiq_t = sum_all_yes
    
    
    if  (sum_all_yes + sum_all_no) == 0: 
        avg_caiq_t=0
    else: 
        avg_caiq_t = round(sum_all_yes[0] / (float(sum_all_yes[0] + sum_all_no[0])),5)
        #print"AVG _ CAIQ T= ", avg_caiq_t
    
#     temp1= ta_assessment * (yes_assessment + no_assessment)    
    total_applicable = 295-sum_all_na[0]
    if  total_applicable==0:
        avg_caiq_c = 1
        #print"Total applicable was Zero so CAIQ_C= ",avg_caiq_c 
    else:
        temp1 = total_applicable * (sum_all_yes[0]+sum_all_no[0])
        #print"temp1 =", temp1
        
    #     temp2 = 2*(ta_assessment - (yes_assessment+no_assessment))
        temp2 = 2*(total_applicable - (sum_all_yes[0]+sum_all_no[0]))
        #print"temp2=", temp2
    #     trust_c= temp1 / (temp2+temp1)
        avg_caiq_c = round(temp1 / float(temp2+temp1),5)
        #print"-------------------------------"
        #print"AVG _ CAIQ_C= ", avg_caiq_c

    avg_caiq_f=0.99
    avg_caiq_e = avg_caiq_t * avg_caiq_c + (1-avg_caiq_c)*avg_caiq_f
    avg_caiq_e = round(avg_caiq_e,5)
    #print"CAIQ AVG _ E Score" , avg_caiq_e
  

    avg_caiq_b = avg_caiq_t * avg_caiq_c # conversion 
    avg_caiq_b = round(avg_caiq_b,5)
    
    avg_caiq_d = (1-avg_caiq_t) * avg_caiq_c
    avg_caiq_d = round(avg_caiq_d,5)
    
    avg_caiq_u = 1-avg_caiq_c 
    avg_caiq_u = round(avg_caiq_u,5)
    
    avg_caiq_a = avg_caiq_f
    
    avg_caiq_w = avg_caiq_b+ avg_caiq_u* avg_caiq_a
    avg_caiq_w = round(avg_caiq_w,5)
    
    #printavg_caiq_a,avg_caiq_b,avg_caiq_c, avg_caiq_d, avg_caiq_e, avg_caiq_f, avg_caiq_t, avg_caiq_u, avg_caiq_w    
    
    misc_data = sum_all_yes,sum_all_no, sum_all_na, sum_all_un, \
                avg_caiq_t, avg_caiq_c,  avg_caiq_f,avg_caiq_e, \
                total_applicable, avg_caiq_b, avg_caiq_d,avg_caiq_u, avg_caiq_w 
                
    #print"i am the CID", cid
    #print"I am misc data", misc_data
                
    cur.execute("""update cprofile set all_yes=:1,all_no=:2,all_na=:3,all_un=:4, 
                    avg_t=:5, avg_c=:6, avg_e=:7,avg_f=:8,
                    avg_b=:9,avg_d=:10,avg_u=:11, avg_w=:12 
                    where cprofile.id=:13""", 
                    (sum_all_yes[0],sum_all_no[0],sum_all_na[0],sum_all_un[0],
                     avg_caiq_t, avg_caiq_c,avg_caiq_e,avg_caiq_f,
                     avg_caiq_b,avg_caiq_d,avg_caiq_u, avg_caiq_w,cid))
    
#     cur.execute("update cprofile set all_yes=111 where cprofile.id=:1",[(cid)]) 
    #print"Rows affected", cur.rowcount
    conn.commit()
    
    return misc_data

#@webapp.route('/composite/<string:cloud_list>')
#def trust_table(cloud_list): # result compare, cloud_detail 
 
def sub_comp_trust(child,pred):   
 # Note: trust values must scale in the range (distrust,uncertainity,trust) or 
    # between (strong distrust, weak distrust, uncertainity, weak trust, strong trust)
    parsed_result = []
    cloud_list = pred, child
    #printcloud_list
    result_compare = caiq_compare(cloud_list).get('result_compare')
    cloud_detail = caiq_compare(cloud_list).get('cloud_detail')
    result_tcp = []
    result_tparent = []
    
    for x in range(len(result_compare)):
        temp1=[]
        for y in range(len(result_compare[x])): 
            if y==0 or y==1:
                temp1.append(result_compare[x][y])
                continue
            temp1.append(parse_trust(result_compare[x][y]))
#             if y==2: 
#                 result_a.append(parse_trust(result_compare[x][y]))       
        
        parsed_result.append(temp1)
        result_tcp.append(tchild_parent(temp1).get('trust_code'))
        
    #print"result tcp", result_tcp.count(1)
    #print"result tcp", result_tcp.count(2)
    #print"result tcp", result_tcp.count(3)
   # pr_trust_b_g_a = pr_trust_bga(result_tcp)
    #pr_trust_a = pr_trust_b_g_a(result_a)
    
    #print"This is trust for child_given_parent", result_tcp
    #print"This is trust for parent", result_tparent
    
    ##print"This is parsed result", parsed_result
    
    return {'parsed_result':parsed_result,'cloud_detail':cloud_detail, 
            #'pr_trust_b_g_a':pr_trust_b_g_a
            }      

def parse_trust(caiq_score):
    # this is the point where a CSP will define its level of trust (May be we will see where the cloud score is itself)  

    if caiq_score>=0 and caiq_score <= 0.4: caiq_score= 'Distrust' 
    elif caiq_score > 0.4 and caiq_score<=0.5: caiq_score='Uncertain'
    elif caiq_score > 0.5 and caiq_score<=1 : caiq_score='Trust'
    
    else: 
        caiq_score="Fake assessment"
    
    sub_trust_value = caiq_score 
    return sub_trust_value

def caiq_compare(cloud_list):
  #  print "clouds to compare ", cloud_list 
    parsed_result = []
    cloud_detail = []
    query_start= 'select ctrustinfo.cgroup_id, caiqcgroup.group_name,' 
    query_end= 'from ctrustinfo INNER JOIN caiqcgroup on ctrustinfo.cgroup_id=caiqcgroup.id group by ctrustinfo.cgroup_id order by ctrustinfo.cgroup_id' 
    inner_query =[]
    ##print"This is cloud list ", cloud_list
    for x in range(len(cloud_list)):
        cloud_detail.append(profile_detail(cloud_list[x]))

        inner_query.append( 'max(case when ctrustinfo.cloud_id ='+str(cloud_list[x])+ ' then ctrustinfo.caiq_e end) as [Peer-'+str(x+1)+']') 
        
    inner_query = ",".join(inner_query)
    full_query = query_start+inner_query+query_end  
    print "This is the full query ", full_query
    
    cur.execute(full_query)
    result_compare = cur.fetchall()


    for x in range(len(result_compare)):
        temp3=[]
        for y in range(len(result_compare[x])): 
            if y==0 or y==1:
                temp3.append(result_compare[x][y])
                continue
            temp3.append(parse_trust(result_compare[x][y]))

        parsed_result.append(temp3) 
    
       
    return {'result_compare':result_compare, 'cloud_detail':cloud_detail, 'parsed_result':parsed_result}

def tchild_parent(list_values):
    
    
    if list_values[2] =='Trust' and list_values[3]=="Trust":  
        return {'trust_code':1} # trust of child given trust of parent 
    elif list_values[2] =='Distrust' and list_values[3]=="Trust":
        return {'trust_code':2} # trust of child given distrust of parent
    elif list_values[2] =='Uncertain' and list_values[3]=="Trust":
        return {'trust_code':3} # trust of child given uncertainity of parent
    else: 
        return {'trust_code':4} # error
    
    
def pr_trust_bga(tcounter):
    
    count_trust = tcounter.count("Trust")
    #printcount_trust
    #printfloat(count_trust)/16 # equal to P(tA|tS) 
    count_distrust = tcounter.count("Distrust") # equal to P(tA|dS) 
    #printfloat(count_distrust)/16  
    
    count_uncertain = tcounter.count("Uncertain") # equal to P(tA|dS) 
    #printfloat(count_uncertain)/16  
    
    return {'count_trust': count_trust, 'count_distrust': count_distrust, 'count_uncertain': count_uncertain}
    
    
def find_or(list_values):
    if list_values[2] =='Trust' or list_values[3]=="Trust":  
        return "Trust"
    else: 
        return "distrust"

@webapp.route('/tpara',methods=['GET', 'POST'])
def trust_settings():
    return render_template("/tpara.html") 

#@webapp.route("/tdgraph")  # trust dependency as graph:: comment for live   
#def td_obj_graph):
def td_obj_graph(nlist,elist):  
    DG=nx.DiGraph()
      

    DG.add_nodes_from(nlist)
    blist=[] # will contain transformed alist as a weighted edge list  
    for myedges in elist: 
        #calculate edge weight as multiple of its node weights 
    #    #print"my edges",myedges[0], myedges[1] 
       # #printDG.node[myedges[0]]['e-score']
        #print"E-score", DG.node[myedges[0]]['obj_trust'][3]
        edge_weight = round((DG.node[myedges[0]]['obj_trust'][3] * DG.node[myedges[1]]['obj_trust'][3]),5)
        my_weighted_edge = (myedges[0], myedges[1], edge_weight)
        blist.append(my_weighted_edge) #  
    #    #print"This is blist",blist 
    DG.add_weighted_edges_from(blist)
#       
    return DG

def caiq_obj_trust(graph):
    
    DG = graph
    ####
    ####  NOW JUST SUM THE EDGES IN A PATH 
    ####
    ####
    
    # get the root node a.k.a home cloud 
    root_nodes= [node for node in DG.nodes() if DG.in_degree(node)==0 and DG.out_degree(node)!=0]
    leaf_nodes = [node for node in DG.nodes() if DG.in_degree(node)!=0 and DG.out_degree(node)==0]
   # #print"root Node is ", root_nodes
 #   #print"leaf nodes are ", leaf_nodes
    
    graph_weight = [] # global trust of this transaction 
    count_paths = []
    tfactor = []
    # traversing all possible paths between a root and its leaves 
    for leaf_node in leaf_nodes:
        #root_nodes= [node for node in DG.nodes() if DG.in_degree(node)==0 and DG.out_degree(node)!=0]
        #printroot_nodes,'-->' , leaf_node
        for root_node in root_nodes:
            sub_graph_weight = 0 # weight for each sub graph having same root same leaf node  
            count_sub_paths = 0  
            #for p in nx.all_shortest_paths(DG,source=root_node,target=leaf_node):
            for p in nx.all_simple_paths(DG,source=root_node,target=leaf_node):     
                #print"I am P", p
                path_weight = 0 # weight for each path 
                for x in range(len(p)-1):
                    path_weight += DG.get_edge_data(p[x],p[x+1])['weight']
                    #printp[x],'-->',p[x+1], path_weight
                avg_path_weight= path_weight/(len(p)-1)
                #printp, '-->', avg_path_weight
                sub_graph_weight += path_weight/(len(p)-1)
                count_sub_paths+=1
            avg_subg_weight = sub_graph_weight/count_sub_paths
            #print"Total paths in this sub graph ", count_sub_paths
            #print"weight for this sub-graph is ", sub_graph_weight
            #print"Average subgraph weight ", avg_subg_weight
        count_paths.append(count_sub_paths)
        #print"Total paths = ", count_paths
        graph_weight.append(avg_subg_weight)
    #printgraph_weight
    #printcount_paths
    
    
    ####
    ### Important note: when the graph has more than one leaves add weight to each path according to the 
    ## ratio of its paths to total paths 
    ## e.g. A graph has two leaves with  two paths - its the average 
    ##    when it has two leaves with one and three paths - make it different !! Think 
    ## number of sub paths div by total paths is the factor of that path 
    ### LOGIC: More the number of paths less reliable OR more the number of nodes in a path less reliable
    ####                
    
    final_trust=0
    for x in range(len(count_paths)): 
        #printfloat(count_paths[x])/sum(count_paths) # total paths in a subgraph / total paths in graph 
        tfactor.append(float(count_paths[x])/sum(count_paths))
        final_trust+= (float(count_paths[x])/sum(count_paths))*graph_weight[x]
        
    #printtfactor,final_trust
    # factorizing ends here 
    
    return {'tfactor':tfactor, 'final_trust':final_trust}


#@webapp.route("/dep_graph/<int:cid>")
@webapp.route("/dep_graph/<int:cid>/<string:trid>")
@webapp.route("/dep_graph/<int:cid>/<string:trid>/<string:engage>")

def dep_graph(cid,trid=0, engage='False'): 
    ##print"this is make graph function "
    #print"Transaction ID=",trid
    #print"Engage =", engage
    nlist = [] # contains the final output 
    elist = []
    fcloud_id = cid
    foreign_cloud = 0
    fcloud_e = 0
    hcloud_id = 0
    hcloud_e=0
    home_cloud=0
    parent_cloud_id = 0
    parent_cloud_e = 0
    parent_cloud = 0
    cur_time = datetime.now()
    gfilename = 0

    if trid=='0':
        cur.execute("select id,cname,avg_t,avg_c,avg_f,avg_e,avg_b,avg_d,avg_u,avg_w, cendpoint from cprofile where login_id=:1", [(session['login_id'])] )
        parent_cloud = home_cloud = cur.fetchone()
        parent_cloud_id = hcloud_id = home_cloud[0]
        
        parent_cloud_t = hcloud_t = home_cloud[2]
        parent_cloud_c = hcloud_c = home_cloud[3]
        parent_cloud_f = hcloud_f = home_cloud[4]
        parent_cloud_e = hcloud_e = home_cloud[5]
        parent_cloud_b = hcloud_b = home_cloud[6]
        parent_cloud_d = hcloud_d = home_cloud[7]
        parent_cloud_u = hcloud_u = home_cloud[8]
        parent_cloud_a = hcloud_a = home_cloud[4]
        parent_cloud_w = hcloud_w = home_cloud[9]
        parent_cloud_ip = hcloud_w = home_cloud[10]
        
# 
# This format is required as json 
# [  ('S',{'obj_trust':(t,c,f,E), 'sub_trust':(b,d,u,a),'dep_trust':(0) } ) , 
#    ('A',{'obj_trust':(t,c,f,E), 'sub_trust':(b,d,u,a),'dep_trust':(0) } )
#  ]
#
        
        nlist.append(( hcloud_id,
                                {'obj_trust':(hcloud_t,hcloud_c,hcloud_f,hcloud_e) , 
                                 'sub_trust':(hcloud_b,hcloud_d,hcloud_u,hcloud_a,hcloud_w),
                                 'dep_trust':(0) 
                                }  
                    ))

        
        cur.execute("select id,cname,avg_t,avg_c,avg_f,avg_e,avg_b,avg_d,avg_u,avg_w,cendpoint from cprofile where id=:1", [(fcloud_id)] )
        foreign_cloud = cur.fetchone()
        fcloud_id = cid 
        fcloud_t = foreign_cloud[2]
        fcloud_c = foreign_cloud[3]
        fcloud_f = foreign_cloud[4]
        fcloud_e = foreign_cloud[5]
        
        fcloud_b = foreign_cloud[6]
        fcloud_d = foreign_cloud[7]
        fcloud_u = foreign_cloud[8]
        fcloud_a = foreign_cloud[4]
        fcloud_w = foreign_cloud[9]
        fcloud_ip = foreign_cloud[10]
        
        nlist.append(( fcloud_id,
                                {'obj_trust':(fcloud_t,fcloud_c,fcloud_f,fcloud_e) , 
                                 'sub_trust':(fcloud_b,fcloud_d,fcloud_u,fcloud_a,fcloud_w),
                                 'dep_trust':(0) 
                                }  
                    ))
        #weight=1
        #myedge = (hcloud_id,fcloud_id,weight)
        myedge = (hcloud_id,fcloud_id)
        elist.append(myedge)
        
        #for items in nlist:
            #print"Final output =", items
        #print"Final edges = ", elist 
        # for objective trust 
        DG = td_obj_graph(nlist, elist)
        caiq_obj_trust_values = caiq_obj_trust(DG) 
        gfilename = 'graphs/'+ str(trid)+'-'+str(hcloud_id)+'-graph.png'
        #print"Control returned. Now its sub trust "
#for subjective trust 
        #return {'sub_graph':subDG, 'sub_comptrust':dep_t_sgraph}
        
        #print"nlist before subjective trust calling", nlist 

        sub_comptrust = subjective_trust((nlist, elist))
        subDG = sub_comptrust['sub_graph']
        subjective_trust_values = sub_comptrust['sub_comptrust']
        sgfilename = 'graphs/'+ str(trid)+'-'+str(hcloud_id)+'-sub-graph.png'

        if engage == 'True':
            
            #print"Starting a new transaction "
            #print"this is add_transaction adding ", hcloud_id, "  +  ", fcloud_id, caiq_obj_trust_values['final_trust']
            # resume here. I am inserting comp-trust values in transactions table 
            
            #first try to start teh comet cloud on remote - if not success dont add trx 
            
            startC4C_info = startC4C(fcloud_ip) 
            if (startC4C_info):
                print "Successfuly started remote host "
            else: 
                exit()
            
            cur.execute("""insert into transactions(hcloud_id,lastpeer, foreignpeers,obj_comptrust,
                b_comp,d_comp,u_comp,a_comp, sub_comptrust,
                status,creationtime, lastactivity,tthreshold) values(?,?,?,?,?,?,?,?,?,?,?,?,?)""", 
                (hcloud_id,fcloud_id,1, caiq_obj_trust_values['final_trust'], 
                 subjective_trust_values[0],subjective_trust_values[1],subjective_trust_values[2],
                 subjective_trust_values[3],subjective_trust_values[4], 710, 
                 cur_time,cur_time,0.75))

            conn.commit()
            cur.execute("select MAX(ID) from transactions")
            trid = cur.fetchone()
            #print"New trx added with id = ", trid[0]

            cur.execute("""insert into subtrans (transaction_id, cloud_id,parent_id,resource_id,timestamp,status)
            values (?,?,?,?,?,?) """, (trid[0],fcloud_id,parent_cloud_id,1,cur_time,710))
            conn.commit()
        
            
            
            ## Transaction dump ... keeps track of all trx for results 
            cur.execute("""insert into trx_dump(trid,hcloud_id,lastpeer, foreignpeers,obj_comptrust,
                b_comp,d_comp,u_comp,a_comp, sub_comptrust,
                status,creationtime, lastactivity,tthreshold) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", 
                (trid[0],hcloud_id,fcloud_id,1, caiq_obj_trust_values['final_trust'], 
                 subjective_trust_values[0],subjective_trust_values[1],subjective_trust_values[2],
                 subjective_trust_values[3],subjective_trust_values[4], 710, 
                 cur_time,cur_time,0.75))
        
            conn.commit()

        
             
            json_fnlist = 'static/trx-json/'+str(trid[0])+'-'+str(hcloud_id)+'-nlist.json'
            json_felist = 'static/trx-json/'+str(trid[0])+'-'+str(hcloud_id)+'-elist.json'

            
            with open(str(json_fnlist), 'w') as fp:
                json.dump(nlist, fp)
            
            with open(str(json_felist), 'w') as fp:
                json.dump(elist, fp)
            
            #print"Just before making a filename for transaction id = ", trid 
            gfilename = 'graphs/'+ str(trid[0])+'-'+str(hcloud_id)+'-graph.png'
            sgfilename = 'graphs/'+ str(trid[0])+'-'+str(hcloud_id)+'-sub-graph.png'
            
            #print"File created after new transaction ", gfilename, sgfilename
            flash ("Started a new transaction ")
           
    elif trid!='0':
        
        cur.execute("select * from transactions where id=?", [(trid)])
        transaction_detail = cur.fetchone()
        hcloud_id = transaction_detail[1]
        
        cur.execute("select id,cname,avg_t,avg_c,avg_f,avg_e,avg_b,avg_d,avg_u,avg_w from cprofile where id=?", [(hcloud_id)])
        home_cloud = cur.fetchone()
        hcloud_id = home_cloud[0]
        hcloud_t = home_cloud[2]
        hcloud_c = home_cloud[3]
        hcloud_f = home_cloud[4]
        hcloud_e = home_cloud[5]
        hcloud_b = home_cloud[6]
        hcloud_d = home_cloud[7]
        hcloud_u = home_cloud[8]
        hcloud_a = home_cloud[4]
        hcloud_w = home_cloud[9]
        
        foreign_peers = transaction_detail[2]
        ##print"Foreign Peers", foreign_peers
        
        cur.execute("select id,cname,avg_t,avg_c,avg_f,avg_e,avg_b,avg_d,avg_u,avg_w from cprofile where login_id=:1", [(session['login_id'])] )
        parent_cloud = cur.fetchone()
        parent_cloud_id = parent_cloud[0]
        parent_cloud_t = parent_cloud[2]
        parent_cloud_c = parent_cloud[3]
        parent_cloud_f = parent_cloud[4]
        parent_cloud_e = parent_cloud[5]
        parent_cloud_b = parent_cloud[6]
        parent_cloud_d = parent_cloud[7]
        parent_cloud_u = parent_cloud[8]
        parent_cloud_a = parent_cloud[4]
        parent_cloud_w = parent_cloud[9]
        
        
        json_fnlist = 'static/trx-json/'+str(trid)+'-'+str(hcloud_id)+'-nlist.json'
        json_felist = 'static/trx-json/'+str(trid)+'-'+str(hcloud_id)+'-elist.json'
        
        with open(str(json_fnlist), 'r') as fp:
            ndata = json.load(fp)
        
        with open(str(json_felist), 'r') as fp:
            edata = json.load(fp)

        nlist = list(ndata)
        elist = list(edata)
        
        
        #print"fresh n list ", nlist 
        
        fcloud_id = cid
        cur.execute("select id,cname,avg_t,avg_c,avg_f,avg_e,avg_b,avg_d,avg_u,avg_w from cprofile where id=:1", [(fcloud_id)] )
        foreign_cloud = cur.fetchone()
        
        fcloud_t = foreign_cloud[2]
        fcloud_c = foreign_cloud[3]
        fcloud_f = foreign_cloud[4]
        fcloud_e = foreign_cloud[5]
        fcloud_b = foreign_cloud[6]
        fcloud_d = foreign_cloud[7]
        fcloud_u = foreign_cloud[8]
        fcloud_a = foreign_cloud[4]
        fcloud_w = foreign_cloud[9]
# 
# This format is required as json 
# [  ('S',{'obj_trust':(t,c,f,E), 'sub_trust':(b,d,u,a),'dep_trust':(0) } ) , 
#    ('A',{'obj_trust':(t,c,f,E), 'sub_trust':(b,d,u,a),'dep_trust':(0) } )
#  ]
#
#         nlist.append(( fcloud_id,
#                                 {u'obj_trust':(fcloud_t,fcloud_c,fcloud_f,fcloud_e) , 
#                                  u'sub_trust':(fcloud_b,fcloud_d,fcloud_u,fcloud_a,fcloud_w),
#                                  u'dep_trust':(0) 
#                                 }  
#                     ))
        
        nlist.append(( fcloud_id,
                                {'obj_trust':(fcloud_t,fcloud_c,fcloud_f,fcloud_e) , 
                                 'sub_trust':(fcloud_b,fcloud_d,fcloud_u,fcloud_a,fcloud_w),
                                 'dep_trust':(0) 
                                }  
                    ))
        
        #print"n list after appending a foreign cloud", nlist 
#        nlist.append([fcloud_id,{ u'e-score':fcloud_e}])

        #elist.append([parent_cloud_id,fcloud_id])
        elist.append((parent_cloud_id,fcloud_id))
        
        DG = td_obj_graph(nlist, elist)
        caiq_obj_trust_values = caiq_obj_trust(DG)
        gfilename = 'graphs/'+ str(trid)+'-'+str(hcloud_id)+'-temp-graph.png'
        
        # subjective here 
        #print"nlist before subjective trust calling", nlist 
        
        sub_comptrust = subjective_trust((nlist, elist))
        subDG = sub_comptrust['sub_graph']
        subjective_trust_values = sub_comptrust['sub_comptrust']
        
        #print'subjective trust' , subjective_trust_values
        
        sgfilename = 'graphs/'+ str(trid[0])+'-'+str(hcloud_id)+'-temp-sub-graph.png'

        if engage == 'True': 
            #print"Updating an old transaction "
            gfilename = 'graphs/'+ str(trid)+'-'+str(hcloud_id)+'-graph.png'
            #print"This is the real name of an updated transaction", gfilename
            with open(str(json_fnlist), 'w') as fp:
                json.dump(nlist, fp)
            with open(str(json_felist), 'w') as fp:
                json.dump(elist, fp)   
            cur_time = datetime.now()
            foreign_peers = foreign_peers+1
            cur.execute("""update transactions set foreignpeers=:1,obj_comptrust=:2,b_comp=:3,d_comp=:4,
            u_comp=:5,a_comp=:6,sub_comptrust=:7,lastpeer=:8,lastactivity=:9 where transactions.ID=:10 """, 
                (foreign_peers, caiq_obj_trust_values['final_trust'], subjective_trust_values[0],subjective_trust_values[1],subjective_trust_values[2],
                 subjective_trust_values[3],subjective_trust_values[4], 
                 fcloud_id,
                 cur_time,
                 trid))

            
            conn.commit()
            
            cur.execute("""insert into subtrans (transaction_id, cloud_id,parent_id,resource_id,timestamp,status)
            values (?,?,?,?,?,?) """, (trid,fcloud_id,parent_cloud_id,1,cur_time,710))
            conn.commit()
            
            ## Transaction dump ... keeps track of all trx for results 
            cur.execute("""insert into trx_dump(trid,hcloud_id,lastpeer, foreignpeers,obj_comptrust,
                b_comp,d_comp,u_comp,a_comp, sub_comptrust,
                status,creationtime, lastactivity,tthreshold) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", 
                (trid,hcloud_id,fcloud_id,foreign_peers, caiq_obj_trust_values['final_trust'], 
                 subjective_trust_values[0],subjective_trust_values[1],subjective_trust_values[2],
                 subjective_trust_values[3],subjective_trust_values[4], 710, 
                 cur_time,cur_time,0.75))
        
            conn.commit()

            
            #print"Transaction is updated "
            gfilename = 'graphs/'+ str(trid)+'-'+str(hcloud_id)+'-graph.png'
            sgfilename = 'graphs/'+ str(trid)+'-'+str(hcloud_id)+'-sub-graph.png'
            #print"Permanent filename is ", gfilename, sgfilename
    
    save_obj_graph(DG,gfilename)
    save_sub_graph(subDG,sgfilename)
    
    
#     print "Home Cloud |Trust Metric | Composite trust | "
#     print home_cloud, "\t", subjective_trust_values, "\t"  
        
    return render_template("graphing.html",
                           engage=engage, gfilename=gfilename,sgfilename=sgfilename, 
                           nlist=nlist,elist=elist, trid=trid,
                           home_cloud=home_cloud,parent_cloud=parent_cloud, foreign_cloud=foreign_cloud,
                           obj_trust=caiq_obj_trust_values, subjective_trust_values=subjective_trust_values,
                           subjective_trust_w=subjective_trust_values,
                           is_xhr= request.is_xhr)

def save_obj_graph(graph,gfilename):
    
    DG = graph
    pos=nx.spring_layout(DG)
    plt.axis('off')
    edge_labels=dict([ ( (u,v,),d['weight'])
                 for u,v,d in DG.edges(data=True)])
    #print"Edge laels ",edge_labels


    nx.draw_networkx_edge_labels(DG,pos,edge_labels=edge_labels)
    nx.draw_networkx(DG,pos, node_size=1500)
    
    full_path = "static/"+ gfilename
    #print"full path=",full_path
    plt.savefig("%s" % full_path)
    plt.clf() 
  
    return 0

def save_sub_graph(graph,sgfilename):
    
    subDG = graph
    
    pos=nx.spring_layout(subDG)
    plt.axis('off')
#     edge_labels=dict( [ ( u,v )
#                 for u,v in subDG.edges(data=True)])
#     edge_labels=dict([ ( (u,v,),d['weight'])
#                  for u,v,d in subDG.edges(data=True)])
#     #print"Edge laels ",edge_labels
        
    #nx.draw_networkx_edge_labels(subDG,pos,edge_labels=edge_labels)
    nx.draw_networkx(subDG,pos, node_size=1500)
    
    full_path = "static/"+ sgfilename
    #print"full path=",full_path
    plt.savefig("%s" % full_path)
    plt.clf() 
  
    return 0
 

#@webapp.route("/testsub")   
def subjective_trust((nlist, elist)):
    
    #print"Check this now ", elist 
    #print"nlist in sub trust", nlist
    
    subDG=nx.DiGraph()    
      
    subDG.add_nodes_from(nlist)
    subDG.add_edges_from(elist)
    #subDG.add_weighted_edges_from(elist)
        
    all_sub_graphs = create_sub_graphs(subDG)['all_sub_graphs']
    #print all_sub_graphs
    
    leaf_nodes = create_sub_graphs(subDG)['leaf_nodes']
    #print"This graph has leaf nodes", leaf_nodes
    
    # nx.set_node_attributes(subDG, 'dep_trust',  ()   )
    for lnode in leaf_nodes:
        subDG.node[lnode]['dep_trust']=subDG.node[lnode]['sub_trust']
       
    
    stack_length = len(all_sub_graphs)
    
    pop_sub_graphs = []
    for item in range (stack_length):
        pop_sub_graphs.append( all_sub_graphs.pop())
    #printpop_sub_graphs
    
    dep_t_sgraph = 0 # holds the dependancy trust values for this sub graph, always n-1
    print pop_sub_graphs
    for sub_graph in pop_sub_graphs: 
        
        for x in range(len(sub_graph)):
        
            pnode = sub_graph[0]
            if x==0:
                continue
            else:
                
             #   print x
                bx = subDG.node[pnode]['sub_trust'][0]
                print"BX", bx, pnode, "BY",sub_graph[x] 
                print"Test ", subDG.node[sub_graph[x]]['dep_trust']
                
                by = subDG.node[sub_graph[x]]['dep_trust'][0]
                  
                dx = subDG.node[pnode]['sub_trust'][1]
                dy = subDG.node[sub_graph[x]]['dep_trust'][1]
                  
                ux = subDG.node[pnode]['sub_trust'][2]
                uy = subDG.node[sub_graph[x]]['dep_trust'][2]
                  
                ax = subDG.node[pnode]['sub_trust'][3]
                ay = subDG.node[sub_graph[x]]['dep_trust'][3]
                
                    # older version- doesnot work for u=0# 
#                 bx_and_by = round(bx*by,5)        
#                 #G[2][3]['weight']
#                 dx_and_dy = round(dx+dy-dx*dy,5)
#                 ux_and_uy = round( (bx * uy) + (ux*by)+(ux*uy), 5 )                 
#                 ax_and_ay = round( (bx*uy*ay + ux*ax*by + ux*ax*uy*ay) / ux_and_uy, 4)
#                
                # r_values = relationship value 
                #subDG[pnode][sub_graph[x]]['r_value'] = (bx_and_by, dx_and_dy, ux_and_uy,ax_and_ay)
                
                
                bx_and_by = round(((bx+ax*ux)*(by+ay*uy) - (1-dx)*(1-dy)*ax*ay)/(1-ax*ay),5)
                
                dx_and_dy = round(dx+dy-dx*dy,5)
                
                ux_and_uy = ((1-dx)*(1-dy)-(bx+ax*ux)*(by+ay*uy))/(1-ax*ay)
                ux_and_uy = round(ux_and_uy,5)
                
                ax_and_ay = round(ax*ay,5) 
                
                wx_and_wy = bx_and_by + ux_and_uy*ax_and_ay
                #print"wx_and_wy",wx_and_wy
                # adding weight to the subDG edge 
                
                #subDG[pnode][sub_graph[x]]['weight'] = (bx_and_by, dx_and_dy, ux_and_uy,ax_and_ay,wx_and_wy)
                
                #printpnode, "and ", sub_graph[x], "==> ",bx_and_by, dx_and_dy, ux_and_uy, ax_and_ay, wx_and_wy
                #consensus here for more than one child 
                # what should be the value of dep trust initial so it be a recursive function 
                if subDG.node[pnode]['dep_trust'] == 0: 
                    subDG.node[pnode]['dep_trust'] =  (bx_and_by, dx_and_dy, ux_and_uy,ax_and_ay,wx_and_wy)
                    
                    #print"Now the dep trust is ", subDG.node[pnode]['dep_trust']
                    dep_t_sgraph=subDG.node[pnode]['dep_trust']
                    
                    
                else: 
                    #print"Consensus is required, dependancy trust is=", subDG.node[pnode]['dep_trust']
                    
                    b_x_A = subDG.node[pnode]['dep_trust'][0]           # already present in dep_trust of pnode 
                    b_x_B = bx_and_by # newly evaluated as B 
                    d_x_A = subDG.node[pnode]['dep_trust'][1]
                    d_x_B = dx_and_dy
                    
                    u_x_A = subDG.node[pnode]['dep_trust'][2]
                    u_x_B = ux_and_uy
                    a_x_A = subDG.node[pnode]['dep_trust'][3]
                    a_x_B = ax_and_ay
                    
                    #print"OLD =",(b_x_A,d_x_A,u_x_A,a_x_A), "@ new= ", (b_x_B,d_x_B,u_x_B,a_x_B) 
                    
                    k= round(u_x_A + u_x_B -(u_x_A*u_x_B),5)
                    #print"This is k", k
                    
                    if k!=0: 
                    
                        b_x_AB = ( b_x_A * u_x_B  +  b_x_B * u_x_A ) / k
                        b_x_AB = round(b_x_AB,5)
                        
                        d_x_AB = ( d_x_A * u_x_B  +  d_x_B * u_x_A ) / k
                        d_x_AB = round(d_x_AB,5)
                        
                        u_x_AB = ( u_x_A * u_x_B ) / k
                        u_x_AB = round(u_x_AB,5)
                        
                        temp = ( a_x_B * u_x_A  +  a_x_A * u_x_B  - ( a_x_A + a_x_B ) * u_x_A * u_x_B )  
                        
                        a_x_AB = temp / ( u_x_A  +  u_x_B - 2 * u_x_A * u_x_B )   
                        a_x_AB = round(a_x_AB,5)
                        
                        # expeced value = b + ua
                        
                    elif k==0: 
                        g=1 # gamma variable as defined by Josang 
                        b_x_AB = (g*b_x_A + b_x_B)/(g+1)
                        b_x_AB = round(b_x_AB,5)
                        
                        d_x_AB = (g*d_x_A + d_x_B)/(g+1)
                        d_x_AB = round(d_x_AB,5)
                        
                        u_x_AB = 0 
                    
                        a_x_AB =  (g*a_x_A + a_x_B)/(g+1)    
                        a_x_AB = round(a_x_AB,5)
                        
                    w_x_AB = b_x_AB + u_x_AB * a_x_AB   
                    w_x_AB = round(w_x_AB,5) 
                    subDG.node[pnode]['dep_trust'] =  (b_x_AB,d_x_AB,u_x_AB,a_x_AB,w_x_AB)
                    #print"After consensus and update node",(pnode), subDG.node[pnode]['dep_trust']
                dep_t_sgraph= subDG.node[pnode]['dep_trust']
                

#     nx.draw_networkx(subDG)
#     plt.savefig("static/graph_sub.png")
#     plt.clf()     
            ##print'dep_t_sgraph', dep_t_sgraph
            
    return {'sub_graph':subDG, 'sub_comptrust':dep_t_sgraph}

# def make_subjective_graph(nlist,elist):  
#     DG=nx.DiGraph()      
# 
#     DG.add_nodes_from(nlist)
#     blist=[] # will contain transformed alist as a weighted edge list  
#     for myedges in elist: 
#     
#         #print"W-score", DG.node[myedges[0]]['sub_trust'][4]
#         edge_weight = round((DG.node[myedges[0]]['obj_trust'][3] * DG.node[myedges[1]]['obj_trust'][3]),4)
#         my_weighted_edge = (myedges[0], myedges[1], edge_weight)
#         blist.append(my_weighted_edge) #  
#     #    #print"This is blist",blist 
#     DG.add_weighted_edges_from(blist)
# #       
#     return DG
#     
#     
#     
#     
#     
#     
#     nlist = nlist 
#     elist = elist
#     
#     my_weighted_edge = (myedges[0], myedges[1], edge_weight)
#     
#     DG.add_nodes_from(nlist)
#     DG.add_edges_from(elist)
# 
#     return DG  
#      
# def create_sub_graphs(graph):
#     
#     DG = nx.DiGraph(graph)
#     root_nodes= [node for node in DG.nodes() if DG.in_degree(node)==0 and DG.out_degree(node)!=0]
#     internal_nodes = [node for node in DG.nodes() if DG.in_degree(node)!=0 and DG.out_degree(node)!=0]
#     leaf_nodes = [node for node in DG.nodes() if DG.in_degree(node)!=0 and DG.out_degree(node)==0]
#                  
#     print"Root Node", root_nodes, "Leaf Node", leaf_nodes, "internal nodes=", internal_nodes
#      
#     all_sub_graphs = []
#     
#     for root_node in root_nodes: 
#         pri_sub_graph = []
#         pri_sub_graph.append(root_node)
#         root_child = DG.successors(root_node)
#         for each_root_child in root_child:
#             pri_sub_graph.append(each_root_child)
#         #print"Pri sub graph = ",pri_sub_graph
#         all_sub_graphs.append(pri_sub_graph)
#         
#     for internal_node in internal_nodes: 
#         
#         internal_sub_graph = []
#         internal_sub_graph.append(internal_node)
#         
#         
#         internal_childs = DG.successors(internal_node)
#         print "internal node+child", internal_node, internal_childs
#         
#         for internal_child in internal_childs: 
#             internal_sub_graph.append(internal_child)
#             
#             
#         all_sub_graphs.append(internal_sub_graph)
#         
#     print "all sub graphs", all_sub_graphs
#     
#     return {
#             'all_sub_graphs':all_sub_graphs, 
#             'root_nodes':root_nodes, 
#             'leaf_nodes':leaf_nodes, 
#             'internal_nodes':internal_nodes
#             } 


def create_sub_graphs(graph):
    
    DG = nx.DiGraph(graph)
    root_nodes= [node for node in DG.nodes() if DG.in_degree(node)==0 and DG.out_degree(node)!=0]
    internal_nodes = [node for node in DG.nodes() if DG.in_degree(node)!=0 and DG.out_degree(node)!=0]
    leaf_nodes = [node for node in DG.nodes() if DG.in_degree(node)!=0 and DG.out_degree(node)==0]
                 
    print"Root Node", root_nodes, "Leaf Node", leaf_nodes, "internal nodes=", internal_nodes
     
    for node in DG.nodes():
        print "I am ", node
    all_sub_graphs = list(nx.dfs_edges(DG, root_nodes[0])) 
     
    
        
    print "all sub graphs", all_sub_graphs
    
    return {
            'all_sub_graphs':all_sub_graphs, 
            'root_nodes':root_nodes, 
            'leaf_nodes':leaf_nodes, 
            'internal_nodes':internal_nodes
            } 

@webapp.route('/biddings', defaults={'trid':'0', 'sort': 'threshold', 'order': 'desc','filter':'RTL'})
@webapp.route('/biddings/<string:trid>')
@webapp.route('/biddings/<string:trid>/<string:sort>')
@webapp.route('/biddings/<string:trid>/<string:sort>/<string:order>')
@webapp.route('/biddings/<string:trid>/<string:sort>/<string:order>/<string:filter>')
def biddings(trid='0', sort='biddings.id', order='asc', filter='RTL'):
    
# select biddings.id,biddings.resource_id, 
# (select resources.rname from resources where resources.id=biddings.resource_id) as rname,
# (select resources.uom from resources where resources.id=biddings.resource_id) as uom ,
# (select resources.rtype from resources where resources.id=biddings.resource_id) as rtype,
# cloud_id,cprofile.cname,cprofile.cendpoint,postedat,expiresat,
# (select statuscodes.description from statuscodes where statuscodes.code=biddings.status) 
# as status,threshold,bidtype from biddings 
# INNER Join cprofile on biddings.cloud_id = cprofile.id  where biddings.bidtype=('RTA') 
# order by resource_id desc
    #print"Transaction ID",trid
    
    query1= '''select biddings.id,cloud_id,cprofile.cname,cprofile.avg_w,cprofile.pvalue,
                (select cprofile.cname
                from transactions inner join cprofile on transactions.hcloud_id=cprofile.id 
                and transactions.id=transaction_id),transaction_id,
                (select foreignpeers from transactions where transactions.id=transaction_id), 
                biddings.resource_id,(select resources.rname from resources 
                where resources.id=biddings.resource_id) as rname,
                biddings.value, (select resources.uom from resources where resources.id=biddings.resource_id) 
                as uom ,
                (select rcategory.cat_name from rcategory INNER join resources on 
                rcategory.id = resources.rtype_id where  resources.id=biddings.resource_id) as rtype,
                strftime('%H:%M:%S',postedat),strftime('%d-%m-%Y',postedat),
                strftime('%H:%M:%S',startsat),strftime('%d-%m-%Y',startsat), 
                strftime('%H:%M:%S',expiresat),strftime('%d-%m-%Y',expiresat),'''
#     strftime('%H:%M:%S',startsat),strftime('%d-%m-%Y',startsat),
#     strftime('%H:%M:%S',expiresat),strftime('%d-%m-%Y',expiresat),''' This is some bug :will fix soon
    
    query2= "(select statuscodes.description from statuscodes where statuscodes.code=biddings.status) as status,threshold,bidtype from biddings" 
    query3 = " INNER Join cprofile on biddings.cloud_id = cprofile.id "
     
    query5= " order by " + sort +' '+ order 
    #if not tfilter== 'all': 
    query4 = " where biddings.bidtype=('" + filter+ "')" 
    
    final_query = query1+query2+query3+query4+query5
   
    #printfinal_query   
    cur.execute(final_query)
    all_biddings = cur.fetchall()
    tcount= len(all_biddings)
    
    
    
#     risk_list = []
#     for each_profile in all_biddings: 
#         
#         temp = each_profile[3] / each_profile[4]
#         risk temp * importance
#         print "This is risk",risk 
    
    return render_template('biddings.html', 
                           tcount=tcount, 
                           biddings=all_biddings,
                # importance = importance, 
                           sort=sort, 
                           order=order,
                           filter=filter, 
                           page='c4crisk',
                           trid = trid,
                           is_xhr=request.is_xhr)


@webapp.route('/edit_transaction', methods=['GET', 'POST'] )
#@webapp.route('/edit/<string:cid>' )
def edit_transaction():
    
    if request.method == 'POST':
        
        if request.form['faction'] == 'New': 
            return redirect('/newtrans')
        
        if request.form['faction'] == 'join':
             
            return redirect("/newsubtrans")
    

    return redirect('/biddings')
 
#@webapp.route("/newtrans", methods=['GET', 'POST'])
# @webapp.route("/newtrans/<string:cid>/<string:trx_type>", methods=['GET', 'POST'])
# def newtrans(cid,trx_type):
# 
#     if request.method == 'GET': 
#         cur.execute("select id,cname,cendpoint,avg_e from cprofile where id=?" , [(cid)]  )
#         whois = cur.fetchone()
#         if whois:
#             #printwhois[0]
#             #printwhois[1]
#             #printwhois[2]
#             #printwhois[3]
#             status = 400
#             curtime = datetime.utcnow()
# 
#             cur.execute("INSERT INTO biddings(cloud_id,threshold,timedate,status,nopeers,type) VALUES (?,?,?,?,?,?)", (whois[0],whois[3],curtime,status,0,trx_type)  ) 
#             conn.commit()
#             message = "Started a new transaction with "+ str(whois[1])+ " as home cloud"
#             flash(message)
# 
#     
#     else: 
#         flash("Error-420")
#     return redirect('/biddings')
        
   # return render_template('newtrans.html', form=form,whois = whois , is_xhr=request.is_xhr )

@webapp.route('/edit_bidding', methods=['GET', 'POST'] )
def edit_bidding():
    
    if request.method == 'POST':
        
        if request.form['bidaction'] == 'Composite':
            flash ("Nothing to do")
            return redirect("/biddings")
        
        if request.form['bidaction'] == 'Add free resource':
            flash ("Supply a resource")
            return redirect('/addrtl')     
        if request.form['bidaction'] == 'Request for resource' :
            flash ("Demand a resource")
            return redirect('/addwta')
#         if request.form['bidaction'] == 'Engage':
#             
#             flash ("Starting a new transaction")
#             hcloud_id = request.form['hcloud_id']
#             fcloud_id = request.form['fcloud_id']
#             #printadd_transaction(hcloud_id,fcloud_id)
#             
#             # This will redirect to show transaction page which is not available now   
#             # Till then it is redirecting to biddings page  
#             #
#             return redirect('/biddings')
        
    return "Hello This is edit bidding"


def add_demand():
    #print"This is demand"
    return "This is demand" 

@webapp.route("/addrtl",methods=['GET', 'POST'])
def addrtl():
    #n = RemoteNode('Alertlogic', '192.168.10.20', 5000)
    ##print'This is n', n.get_id()
    #current_s = n.get_service()
    ##printcurrent_s.get_disks()
    
    form = AddResource(request.form)
#     cur.execute("select id, rname from resources")
#     resource_choices= cur.fetchall()
#     #printresource_choices
    
    form.set_choices()
    #storage =  round ( current_service.get_disks()[0]['space_free']/float((1000*1000*1000)),2)

    #storage =  round ( current_s.get_disks()[0]['space_free']/float((1000*1000*1000)),2)
    if request.method == 'POST':
        
        cloud_id = request.form['cloud_id']
        cname = request.form['cname']
        rtype = request.form['rtype']
        resource = request.form['resource']
        rvalue = request.form['rvalue']
       # rthreshold = request.form['rthreshold']
        uom =  request.form['uom']
        rqty = request.form['rqty']
        lstart =  request.form['lstart']
        expiresat =  request.form['expiresat']
        accept_tos = request.form['accept_tos']
        cur_time = datetime.now()
        
        if form.validate():
            
            cur.execute( 'insert into biddings(bidtype,cloud_id,postedat,startsat,expiresat,status,threshold,resource_id,value,quantity) values (?,?,?,?,?,?,?,?,?,?)', ('RTL',cloud_id,cur_time,lstart,expiresat,600,0.90,1,rvalue,rqty))  
            conn.commit()
            return redirect( request.referrer)
        else:
            
            flash('All the form fields are required. ')
        
    return render_template('addresource.html', form=form,is_xhr=request.is_xhr )
    
    
@webapp.route('/addwta', defaults={'trid': 0}, methods=['GET', 'POST'])
@webapp.route("/addwta/<string:trid>", methods=['GET', 'POST'])
def addwta(trid):
    form = AddResource(request.form)
    form.set_choices()
#     cur.execute("select id, rname from resources")
#     resource_choices= cur.fetchall()
#     #printresource_choices
    
#     if request.method == 'GET':
    #print"This is tr id ", trid
    cur.execute("select id,(select cname from cprofile where cprofile.id = transactions.hcloud_id), hcloud_id from transactions where id=?", [(trid)])
    transaction_data= cur.fetchone()
    #printtransaction_data
        
    if request.method == 'POST':  
        transaction_id = request.form['transaction_id']
        hcloud_id = request.form['hcloud_id']
        hcloud_name = request.form['hcloud_name']
        cloud_id = request.form['cloud_id']
        cname = request.form['cname']
        rtype = request.form['rtype']
        resource = request.form['resource']
        rvalue = request.form['rvalue']
        uom =  request.form['uom']
        rqty = request.form['rqty']
        lstart =  request.form['lstart']
        expiresat =   request.form['expiresat']
        accept_tos = request.form['accept_tos']
        cur_time = datetime.now()
        
        if form.validate():
            cur.execute( 'insert into biddings(bidtype,cloud_id,transaction_id,postedat,startsat,expiresat,status,threshold,resource_id,value,quantity) values (?,?,?,?,?,?,?,?,?,?,?)', ('WTA',cloud_id,transaction_id,cur_time,lstart,expiresat,600,0.90,1,rvalue,rqty))  
            conn.commit()
        
            return redirect( "/biddings")
        else:
            
            flash('All the form fields are required. ')
        
    #return render_template('rfr.html', transaction_data=transaction_data,form=form,is_xhr=request.is_xhr )
    return render_template('search.html', transaction_data=transaction_data,form=form,is_xhr=request.is_xhr )



# @webapp.route("/addtransaction", methods=["GET", "POST"])      
# def add_transaction():
#     
#     if request.form['bidaction'] == 'Engage':
#              
#         flash ("Starting a new transaction")
#         hcloud_id = request.form['hcloud_id']
#         fcloud_id = request.form['fcloud_id']
#         comp_trust = request.form['comp_trust']
#         
#         #print"this is add_transaction adding ", hcloud_id, "  +  ", fcloud_id, comp_trust
#         cur.execute("insert into transactions(hcloud_id,lastpeer, foreignpeers,obj_obj_comptrust) values(?,?,?,?)", (hcloud_id,fcloud_id,1,comp_trust))
#         conn.commit()
#     
#     #cur.execute("select * from transactions where "
#     
#     # This will redirect to show transaction page which is not available now   
#         # Till then it is redirecting to biddings page  
#         #
#         
#     return redirect("/biddings")

    
@webapp.route('/transactions', defaults={'sort': 'obj_comptrust', 'order': 'desc'})
@webapp.route('/transactions/<string:sort>')
@webapp.route('/transactions/<string:sort>/<string:order>')
def transactions(sort='id', order='asc'):
    
    
    query1= """select ID,hcloud_id, (select cname from cprofile where id=hcloud_id) as 'homecloud', 
             (select cname from cprofile where id=lastpeer) as 'lastpeer', foreignpeers, 
             strftime('%d-%m-%Y',creationtime),strftime('%H:%M:%S',creationtime),
             strftime('%d-%m-%Y',lastactivity),strftime('%H:%M:%S',lastactivity),
             tthreshold, obj_comptrust,   
             (select statuscodes.description from statuscodes where statuscodes.code=transactions.status), 
             sub_comptrust 
              """  
    query2 = "from transactions  "   
    query3 =  "where hcloud_id=%s" % (session['reg_clouds'][0][0]) 
    query4 = " order by " + sort + ' ' + order 
    
    ##printquery1, query2 , query3, query4
    cur.execute (query1+query2+query3+query4)
    transactions = cur.fetchall()
    #print(session['reg_clouds'][0][0])
    
    query11= """select subtrans.transaction_id, 
                transactions.hcloud_id, 
                (select cname from cprofile where id =transactions.hcloud_id) as 'Home cloud', 
                     
                subtrans.parent_id, (select cname from cprofile where id =parent_id) as 'Parent Name',
                transactions.foreignpeers, 
                transactions.creationtime, transactions.lastactivity, 
                transactions.tthreshold, transactions.obj_comptrust, 
                (select description from statuscodes where code = subtrans.status) as 'Status', 
                transactions.sub_comptrust                  
                from subtrans 
                inner join transactions on transactions.ID=subtrans.transaction_id where cloud_id=%s""" % (session['reg_clouds'][0][0]) 
        
    cur.execute(query11)
    subtrans = cur.fetchall()
    
    
    return render_template('transactions.html', transactions=transactions,subtrans=subtrans, 
                            sort=sort, order=order, page='transactions', is_xhr=request.is_xhr)
    
@webapp.route('/transaction/<int:trid>', defaults={'section': 'overview'})
@webapp.route('/transaction/<string:triid>', defaults={'section': 'overview'})
@webapp.route('/transaction/<int:trid>/<string:section>')
def transaction(trid, section):
    ##print"This is profile function", cid
    valid_sections = [
        'overview',
        'objective',
        'subjective',
        ]

    if section not in valid_sections:
        errmsg = 'Invalid subsection when trying to view detail' 
        return render_template('error.html', error=errmsg), 404
    
    
    query11= """select subtrans.transaction_id, 
            transactions.hcloud_id, (select cname from cprofile where id =transactions.hcloud_id) as 'Home cloud',
            subtrans.cloud_id, (select cname from cprofile where id =cloud_id) as 'My Name', 
            subtrans.parent_id, (select cname from cprofile where id =parent_id) as 'Parent Name',
            subtrans.timestamp, 
            subtrans.resource_id, (select resources.rname from resources where resources.id=subtrans.resource_id) as 'Resource', 
            transactions.creationtime, transactions.lastactivity, transactions.tthreshold, transactions.obj_comptrust,transactions.sub_comptrust, 
            subtrans.status, (select description from statuscodes where code = subtrans.status) as 'Status',                
    
            (select avg_e from cprofile where id =transactions.hcloud_id) as 'Home CAIQ',
            (select avg_e from cprofile where id =cloud_id) as 'My CAIQ',
            (select avg_e from cprofile where id =parent_id) as 'Parent CAIQ', 
            transactions.b_comp,transactions.d_comp,transactions.u_comp,transactions.a_comp
            
            
            from subtrans inner join transactions on transactions.ID=subtrans.transaction_id where transactions.ID=%s""" % (trid)
    cur.execute(query11)
    transaction_data = cur.fetchall()
    
    #printtransaction_data
    
    gfilename = 'graphs/'+ str(trid)+'-'+str(transaction_data[0][1])+'-graph.png'
    sgfilename = 'graphs/'+ str(trid)+'-'+str(transaction_data[0][1])+'-sub-graph.png'
    #printgfilename, sgfilename
    
    context = {
        'transaction_data' : transaction_data, 
        'section': section,
        'gfilename': gfilename,
        'sgfilename': sgfilename,
        'page': 'transactions',
        'is_xhr': request.is_xhr
    }
    
    if section ==  'objective':

        context['objective'] = 'This is objective trust'
       
        
    elif section == 'subjective':
        context['subjective'] = 'This is subjective trust'
       
    return render_template(
        'transaction/%s.html' % section,
        **context
    ) 
    
    